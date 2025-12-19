from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from inventario.models import Insumo
from caja.models import Venta, DetalleVenta
from clinica.models import Consulta, Hospitalizacion
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal
from functools import wraps


def validar_fecha(fecha_str):
    """Valida que una cadena sea una fecha válida en formato YYYY-MM-DD"""
    if not fecha_str or not fecha_str.strip():
        return ''
    try:
        datetime.strptime(fecha_str.strip(), '%Y-%m-%d')
        return fecha_str.strip()
    except ValueError:
        return ''


def admin_required(view_func):
    """Decorador que verifica si el usuario es admin (staff/superuser/rol administracion)"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Verificar si el usuario está autenticado
        if not request.user.is_authenticated:
            return redirect('login')
        
        # Verificar si es admin/staff/superuser
        if request.user.is_staff or request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        
        # Verificar si tiene rol de administración
        if hasattr(request.user, 'rol') and request.user.rol == 'administracion':
            return view_func(request, *args, **kwargs)
        
        # Si no es admin, mostrar error 403
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('No tiene permiso para acceder a este recurso.')
    
    return wrapper


@admin_required
def index(request):
    return render(request, 'reportes/index.html')


@admin_required
def reporte_financieros(request):
    """Vista para reportes financieros (ventas/caja) con sesiones expandibles"""
    from django.db.models import Sum, Q
    from caja.models import SesionCaja, Venta
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Obtener filtros
    fecha_desde = validar_fecha(request.GET.get('fecha_desde', ''))
    fecha_hasta = validar_fecha(request.GET.get('fecha_hasta', ''))
    metodo_pago = request.GET.get('metodo_pago', '').strip()
    
    # Query de sesiones de caja
    sesiones = SesionCaja.objects.all().prefetch_related('ventas')
    
    # Filtro por fecha desde
    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        sesiones = sesiones.filter(fecha_apertura__date__gte=fecha_desde_dt.date())
    
    # Filtro por fecha hasta
    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        # Sumar 1 día para incluir todo el día hasta
        fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
        sesiones = sesiones.filter(fecha_apertura__date__lt=fecha_hasta_dt.date())
    
    # Procesar cada sesión con sus ventas filtradas
    sesiones_data = []
    total_periodo = Decimal('0.00')
    
    for sesion in sesiones:
        # Obtener todas las ventas pagadas de la sesión
        ventas = sesion.ventas.filter(estado='pagado')
        
        # Aplicar filtros a las ventas si se especificaron
        if metodo_pago:
            ventas = ventas.filter(metodo_pago=metodo_pago)
        
        # Filtrar por fecha de pago si se especificó
        if fecha_desde:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            ventas = ventas.filter(fecha_pago__date__gte=fecha_desde_dt.date())
        
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            ventas = ventas.filter(fecha_pago__date__lt=fecha_hasta_dt.date())
        
        # Solo incluir sesiones que tengan ventas después de aplicar filtros
        if ventas.exists():
            total_sesion = ventas.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
            total_periodo += total_sesion
            
            sesiones_data.append({
                'sesion': sesion,
                'ventas': ventas.order_by('-fecha_pago'),
                'total_sesion': total_sesion,
            })
    
    # Contexto para el template
    context = {
        'sesiones': sesiones_data,
        'total_periodo': total_periodo,
    }
    return render(request, 'reportes/financieros.html', context)


@admin_required
def exportar_financieros_excel(request):
    """Exporta el reporte financiero a Excel con sesiones y ventas"""
    from django.db.models import Sum
    from caja.models import SesionCaja, Venta
    from django.utils import timezone
    from datetime import datetime, timedelta
    import io
    
    # Obtener filtros (mismos que en reporte_financieros)
    fecha_desde = validar_fecha(request.GET.get('fecha_desde', ''))
    fecha_hasta = validar_fecha(request.GET.get('fecha_hasta', ''))
    metodo_pago = request.GET.get('metodo_pago', '').strip()
    
    # Query de sesiones de caja
    sesiones = SesionCaja.objects.all().prefetch_related('ventas')
    
    # Filtro por fecha desde
    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        sesiones = sesiones.filter(fecha_apertura__date__gte=fecha_desde_dt.date())
    
    # Filtro por fecha hasta
    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
        sesiones = sesiones.filter(fecha_apertura__date__lt=fecha_hasta_dt.date())
    
    # Procesar cada sesión con sus ventas filtradas
    sesiones_data = []
    total_periodo = Decimal('0.00')
    
    for sesion in sesiones:
        # Obtener todas las ventas pagadas de la sesión
        ventas = sesion.ventas.filter(estado='pagado')
        
        # Aplicar filtros a las ventas si se especificaron
        if metodo_pago:
            ventas = ventas.filter(metodo_pago=metodo_pago)
        
        # Filtrar por fecha de pago si se especificó
        if fecha_desde:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            ventas = ventas.filter(fecha_pago__date__gte=fecha_desde_dt.date())
        
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            ventas = ventas.filter(fecha_pago__date__lt=fecha_hasta_dt.date())
        
        # Solo incluir sesiones que tengan ventas después de aplicar filtros
        if ventas.exists():
            total_sesion = ventas.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
            total_periodo += total_sesion
            
            sesiones_data.append({
                'sesion': sesion,
                'ventas': ventas.order_by('-fecha_pago'),
                'total_sesion': total_sesion,
            })
    
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes Financieros"
    
    # Encabezado del reporte
    ws.append(['REPORTE FINANCIERO - SESIONES DE CAJA'])
    ws.append(['Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M')])
    ws.append(['Generado por:', request.user.nombre if hasattr(request.user, 'nombre') else request.user.username])
    ws.append([])  # Línea en blanco
    
    # Mostrar filtros aplicados
    ws.append(['FILTROS APLICADOS:'])
    if fecha_desde:
        ws.append(['Fecha desde:', fecha_desde])
    if fecha_hasta:
        ws.append(['Fecha hasta:', fecha_hasta])
    if metodo_pago:
        metodo_display = dict(Venta.METODO_PAGO_CHOICES).get(metodo_pago, metodo_pago)
        ws.append(['Método de pago:', metodo_display])
    ws.append([])  # Línea en blanco
    ws.append(['Total de sesiones:', len(sesiones_data)])
    ws.append(['Total del período:', f'${total_periodo:,.2f}'])
    ws.append([])  # Línea en blanco
    
    # Tablas de sesiones y ventas
    for idx, item in enumerate(sesiones_data, 1):
        sesion = item['sesion']
        ventas = item['ventas']
        total_sesion = item['total_sesion']
        
        # Encabezado de sesión
        ws.append([f'SESIÓN {idx}'])
        ws.append(['Fecha Apertura:', sesion.fecha_apertura.strftime('%d/%m/%Y %H:%M')])
        ws.append(['Usuario Apertura:', f"{sesion.usuario_apertura.nombre} {sesion.usuario_apertura.apellido}"])
        
        if sesion.fecha_cierre:
            ws.append(['Fecha Cierre:', sesion.fecha_cierre.strftime('%d/%m/%Y %H:%M')])
        if sesion.usuario_cierre:
            ws.append(['Usuario Cierre:', f"{sesion.usuario_cierre.nombre} {sesion.usuario_cierre.apellido}"])
        
        estado_sesion = 'Cerrada' if sesion.esta_cerrada else 'Abierta'
        ws.append(['Estado:', estado_sesion])
        ws.append([])  # Línea en blanco
        
        # Tabla de ventas de esta sesión
        headers_venta = ['Fecha/Hora', 'Nº Venta', 'Paciente', 'Método de Pago', 'Origen', 'Total']
        ws.append(headers_venta)
        
        for venta in ventas:
            paciente_nombre = venta.paciente.nombre if venta.paciente else '-'
            metodo_display = dict(Venta.METODO_PAGO_CHOICES).get(venta.metodo_pago, venta.metodo_pago) if venta.metodo_pago else '-'
            
            # Mapear tipo_origen a nombre legible
            tipo_origen_map = {
                'consulta': 'Consulta',
                'hospitalizacion': 'Hospitalización',
                'venta_libre': 'Venta Libre'
            }
            tipo_origen_display = tipo_origen_map.get(venta.tipo_origen, venta.tipo_origen)
            
            ws.append([
                venta.fecha_pago.strftime('%d/%m/%Y %H:%M') if venta.fecha_pago else '-',
                venta.numero_venta,
                paciente_nombre,
                metodo_display,
                tipo_origen_display,
                float(venta.total)
            ])
        
        # Subtotal de sesión
        ws.append(['', '', '', 'Subtotal sesión:', '', float(total_sesion)])
        ws.append([])  # Línea en blanco
    
    # Total del período
    ws.append(['', '', '', 'TOTAL DEL PERÍODO:', '', float(total_periodo)])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Generar respuesta HTTP
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f'reporte_financiero_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        excel_buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@admin_required
def reporte_inventario(request):
    from django.db.models import Count, Q
    
    # Query base
    insumos = Insumo.objects.all()
    
    # Filtro de marca
    marca = request.GET.get('marca', '').strip()
    if marca:
        insumos = insumos.filter(marca=marca)
    
    # Filtro de estado
    estado = request.GET.get('estado', '').strip()
    if estado == 'activo':
        insumos = insumos.filter(archivado=False)
    elif estado == 'archivado':
        insumos = insumos.filter(archivado=True)
    
    # Filtros de stock
    stock_min = request.GET.get('stock_min', '').strip()
    if stock_min:
        try:
            insumos = insumos.filter(stock_actual__gte=int(stock_min))
        except (ValueError, TypeError):
            pass
    
    stock_max = request.GET.get('stock_max', '').strip()
    if stock_max:
        try:
            insumos = insumos.filter(stock_actual__lte=int(stock_max))
        except (ValueError, TypeError):
            pass
    
    # Filtros de precio
    precio_min = request.GET.get('precio_min', '').strip()
    if precio_min:
        try:
            insumos = insumos.filter(precio_venta__gte=float(precio_min))
        except (ValueError, TypeError):
            pass
    
    precio_max = request.GET.get('precio_max', '').strip()
    if precio_max:
        try:
            insumos = insumos.filter(precio_venta__lte=float(precio_max))
        except (ValueError, TypeError):
            pass
    
    # Filtros avanzados basados en ventas
    vendidos = request.GET.get('vendidos', '').strip()
    en_consultas = request.GET.get('en_consultas', '').strip()
    en_hospitalizaciones = request.GET.get('en_hospitalizaciones', '').strip()
    fecha_desde = validar_fecha(request.GET.get('fecha_desde', ''))
    fecha_hasta = validar_fecha(request.GET.get('fecha_hasta', ''))
    
    # Filtros de uso en consultas - LÓGICA SIMPLE
    if en_consultas == 'si':
        # Mostrar SOLO los que fueron usados en consultas
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='consulta',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        insumos = insumos.filter(idInventario__in=ids_usados) if ids_usados else insumos.none()
    elif en_consultas == 'no':
        # Excluir los que fueron usados en consultas
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='consulta',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        if ids_usados:
            insumos = insumos.exclude(idInventario__in=ids_usados)
    
    # Filtros de uso en hospitalizaciones - LÓGICA SIMPLE
    if en_hospitalizaciones == 'si':
        # Mostrar SOLO los que fueron usados en hospitalizaciones
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='hospitalizacion',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        insumos = insumos.filter(idInventario__in=ids_usados) if ids_usados else insumos.none()
    elif en_hospitalizaciones == 'no':
        # Excluir los que fueron usados en hospitalizaciones
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='hospitalizacion',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        if ids_usados:
            insumos = insumos.exclude(idInventario__in=ids_usados)
    
    # Construir filtro para anotación de veces vendido
    filtro_count = Q()
    if fecha_desde:
        filtro_count &= Q(detalleventa__venta__fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        filtro_count &= Q(detalleventa__venta__fecha_creacion__date__lte=fecha_hasta)
    
    # Anotar veces vendido
    insumos = insumos.annotate(veces_vendido=Count('detalleventa', filter=filtro_count)).order_by('medicamento')
    
    # Aplicar filtro vendidos DESPUÉS de la anotación
    if vendidos == 'si':
        insumos = insumos.filter(veces_vendido__gt=0)
    elif vendidos == 'no':
        insumos = insumos.filter(veces_vendido=0)
    
    # Obtener marcas únicas para el filtro
    marcas = Insumo.objects.exclude(marca__isnull=True).exclude(marca='').values_list('marca', flat=True).distinct().order_by('marca')
    
    context = {
        'insumos': insumos,
        'marcas': marcas,
        'url_exportar': reverse('reportes:exportar_inventario_excel')
    }
    return render(request, 'reportes/inventario.html', context)


@admin_required
def exportar_inventario_excel(request):
    from django.db.models import Count, Q
    
    # Aplicar los mismos filtros que en la vista
    insumos = Insumo.objects.all()
    
    # Filtro de marca
    marca = request.GET.get('marca', '').strip()
    if marca:
        insumos = insumos.filter(marca=marca)
    
    # Filtro de estado
    estado = request.GET.get('estado', '').strip()
    if estado == 'activo':
        insumos = insumos.filter(archivado=False)
    elif estado == 'archivado':
        insumos = insumos.filter(archivado=True)
    
    # Filtros de stock
    stock_min = request.GET.get('stock_min', '').strip()
    if stock_min:
        try:
            insumos = insumos.filter(stock_actual__gte=int(stock_min))
        except (ValueError, TypeError):
            pass
    
    stock_max = request.GET.get('stock_max', '').strip()
    if stock_max:
        try:
            insumos = insumos.filter(stock_actual__lte=int(stock_max))
        except (ValueError, TypeError):
            pass
    
    # Filtros de precio
    precio_min = request.GET.get('precio_min', '').strip()
    if precio_min:
        try:
            insumos = insumos.filter(precio_venta__gte=float(precio_min))
        except (ValueError, TypeError):
            pass
    
    precio_max = request.GET.get('precio_max', '').strip()
    if precio_max:
        try:
            insumos = insumos.filter(precio_venta__lte=float(precio_max))
        except (ValueError, TypeError):
            pass
    
    # Filtros avanzados basados en ventas
    vendidos = request.GET.get('vendidos', '').strip()
    en_consultas = request.GET.get('en_consultas', '').strip()
    en_hospitalizaciones = request.GET.get('en_hospitalizaciones', '').strip()
    fecha_desde = validar_fecha(request.GET.get('fecha_desde', ''))
    fecha_hasta = validar_fecha(request.GET.get('fecha_hasta', ''))
    
    # Filtros de uso en consultas - LÓGICA SIMPLE
    if en_consultas == 'si':
        # Mostrar SOLO los que fueron usados en consultas
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='consulta',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        insumos = insumos.filter(idInventario__in=ids_usados) if ids_usados else insumos.none()
    elif en_consultas == 'no':
        # Excluir los que fueron usados en consultas
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='consulta',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        if ids_usados:
            insumos = insumos.exclude(idInventario__in=ids_usados)
    
    # Filtros de uso en hospitalizaciones - LÓGICA SIMPLE
    if en_hospitalizaciones == 'si':
        # Mostrar SOLO los que fueron usados en hospitalizaciones
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='hospitalizacion',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        insumos = insumos.filter(idInventario__in=ids_usados) if ids_usados else insumos.none()
    elif en_hospitalizaciones == 'no':
        # Excluir los que fueron usados en hospitalizaciones
        ids_usados = list(DetalleVenta.objects.filter(
            tipo='insumo',
            venta__tipo_origen='hospitalizacion',
            insumo_id__isnull=False
        ).values_list('insumo_id', flat=True).distinct())
        if ids_usados:
            insumos = insumos.exclude(idInventario__in=ids_usados)
    
    # Construir filtro para anotación de veces vendido
    filtro_count = Q()
    if fecha_desde:
        filtro_count &= Q(detalleventa__venta__fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        filtro_count &= Q(detalleventa__venta__fecha_creacion__date__lte=fecha_hasta)
    
    # Anotar veces vendido
    insumos = insumos.annotate(veces_vendido=Count('detalleventa', filter=filtro_count)).order_by('medicamento')
    
    # Aplicar filtro vendidos DESPUÉS de la anotación
    if vendidos == 'si':
        insumos = insumos.filter(veces_vendido__gt=0)
    elif vendidos == 'no':
        insumos = insumos.filter(veces_vendido=0)
    
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Agregar información de filtros aplicados en las primeras filas
    ws.append(['REPORTE DE INVENTARIO'])
    ws.append(['Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M')])
    ws.append(['Generado por:', request.user.nombre if hasattr(request.user, 'nombre') else request.user.username])
    ws.append([])  # Línea en blanco
    
    # Mostrar filtros aplicados
    ws.append(['FILTROS APLICADOS:'])
    if marca:
        ws.append(['Marca:', marca])
    if estado:
        ws.append(['Estado:', 'Activo' if estado == 'activo' else 'Archivado'])
    if stock_min:
        ws.append(['Stock mínimo:', stock_min])
    if stock_max:
        ws.append(['Stock máximo:', stock_max])
    if precio_min:
        ws.append(['Precio mínimo:', f'${precio_min}'])
    if precio_max:
        ws.append(['Precio máximo:', f'${precio_max}'])
    if fecha_desde:
        ws.append(['Fecha desde:', fecha_desde])
    if fecha_hasta:
        ws.append(['Fecha hasta:', fecha_hasta])
    if vendidos:
        ws.append(['Vendidos:', 'Solo vendidos' if vendidos == 'si' else 'No vendidos'])
    if en_consultas:
        ws.append(['Usados en consultas:', 'Sí' if en_consultas == 'si' else 'No'])
    if en_hospitalizaciones:
        ws.append(['Usados en hospitalizaciones:', 'Sí' if en_hospitalizaciones == 'si' else 'No'])
    
    ws.append([])  # Línea en blanco
    ws.append(['Total de registros:', insumos.count()])
    ws.append([])  # Línea en blanco
    
    # Encabezados de datos
    headers = ['Medicamento', 'Marca', 'Stock Actual', 'Precio Venta', 'Veces Vendido', 'Estado']
    ws.append(headers)
    
    # Agregar datos
    for insumo in insumos:
        estado_txt = 'Archivado' if insumo.archivado else 'Activo'
        marca_txt = insumo.marca if insumo.marca else '-'
        precio = float(insumo.precio_venta) if insumo.precio_venta else 0.0
        
        ws.append([
            insumo.medicamento,
            marca_txt,
            float(insumo.stock_actual),
            precio,
            insumo.veces_vendido,
            estado_txt
        ])
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Guardar los filtros aplicados para registro
    filtros_dict = {
        'marca': marca,
        'estado': estado,
        'stock_min': stock_min,
        'stock_max': stock_max,
        'precio_min': precio_min,
        'precio_max': precio_max,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'vendidos': vendidos,
        'en_consultas': en_consultas,
        'en_hospitalizaciones': en_hospitalizaciones,
    }
    
    # Generar resumen legible de filtros
    resumen_parts = []
    if marca:
        resumen_parts.append(f"Marca: {marca}")
    if estado:
        resumen_parts.append(f"Estado: {dict([('disponible','Disponible'),('no_disponible','No Disponible')]).get(estado, estado)}")
    if stock_min:
        resumen_parts.append(f"Stock mínimo: {stock_min}")
    if stock_max:
        resumen_parts.append(f"Stock máximo: {stock_max}")
    if precio_min:
        resumen_parts.append(f"Precio desde: ${precio_min}")
    if precio_max:
        resumen_parts.append(f"Precio hasta: ${precio_max}")
    if fecha_desde:
        resumen_parts.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        resumen_parts.append(f"Hasta: {fecha_hasta}")
    if vendidos:
        resumen_parts.append(f"Vendidos: {dict([('si','Sí'),('no','No')]).get(vendidos, vendidos)}")
    if en_consultas:
        resumen_parts.append(f"En consultas: {dict([('si','Sí'),('no','No')]).get(en_consultas, en_consultas)}")
    if en_hospitalizaciones:
        resumen_parts.append(f"En hospitalizaciones: {dict([('si','Sí'),('no','No')]).get(en_hospitalizaciones, en_hospitalizaciones)}")
    
    resumen_filtros = " | ".join(resumen_parts) if resumen_parts else "Sin filtros aplicados"
    
    # Generar resumen de resultados (tipo tabla de texto)
    resultados_texto = []
    resultados_texto.append(f"{'MEDICAMENTO':<40} {'MARCA':<20} {'STOCK':<10} {'PRECIO':<15} {'VENDIDO':<10} {'ESTADO':<15}")
    resultados_texto.append("=" * 110)
    
    for insumo in insumos[:100]:  # Limitar a 100 registros para no saturar el campo
        veces_vendido = DetalleVenta.objects.filter(
            insumo=insumo,
            venta__estado='pagado'
        ).count()
        
        medicamento = insumo.medicamento[:38] if len(insumo.medicamento) > 38 else insumo.medicamento
        marca = (insumo.marca[:18] if len(insumo.marca) > 18 else insumo.marca) if insumo.marca else "-"
        stock = str(insumo.stock_actual)
        precio = f"${insumo.precio_venta:,.2f}" if insumo.precio_venta else "$0.00"
        vendido = str(veces_vendido)
        estado = "Activo" if not insumo.archivado else "Archivado"
        
        resultados_texto.append(f"{medicamento:<40} {marca:<20} {stock:<10} {precio:<15} {vendido:<10} {estado:<15}")
    
    if insumos.count() > 100:
        resultados_texto.append(f"\n... y {insumos.count() - 100} registros más")
    
    observaciones_completas = f"""FILTROS APLICADOS:
{resumen_filtros}

RESULTADOS ({insumos.count()} registros):
{chr(10).join(resultados_texto)}

Generado desde IP: {request.META.get('REMOTE_ADDR', 'IP no disponible')}"""
    
    # Crear registro del reporte (sin guardar archivo físico - Render usa sistema efímero)
    from .models import ReporteGenerado
    
    ReporteGenerado.objects.create(
        tipo='inventario',
        usuario=request.user,
        filtros=filtros_dict,
        resumen_filtros=resumen_filtros,
        total_registros=insumos.count(),
        observaciones=observaciones_completas
    )
    
    # Preparar respuesta HTTP
    import io
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f'reporte_inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        excel_buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@admin_required
def reporte_caja(request):
    ventas = Venta.objects.select_related(
        'usuario_cobro', 'usuario_creacion', 'paciente'
    ).order_by('-fecha_creacion')
    context = {
        'ventas': ventas,
        'url_exportar': reverse('reportes:exportar_caja_excel')
    }
    return render(request, 'reportes/caja.html', context)


@admin_required
def exportar_caja_excel(request):
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Caja"
    
    # Encabezados
    headers = ['Fecha', 'Número Venta', 'Total', 'Estado', 'Usuario']
    ws.append(headers)
    
    # Obtener datos
    ventas = Venta.objects.select_related(
        'usuario_cobro', 'usuario_creacion', 'paciente'
    ).order_by('-fecha_creacion')
    
    # Agregar datos
    for venta in ventas:
        fecha = venta.fecha_creacion.strftime('%d/%m/%Y %H:%M') if venta.fecha_creacion else '-'
        usuario = venta.usuario_cobro.get_full_name() if venta.usuario_cobro else venta.usuario_creacion.get_full_name()
        total = float(venta.total) if venta.total else 0.0
        estado = venta.get_estado_display()
        
        ws.append([
            fecha,
            venta.numero_venta,
            total,
            estado,
            usuario
        ])
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_caja.xlsx'
    
    # Guardar workbook en response
    wb.save(response)
    
    return response


@admin_required
def reporte_clinica(request):
    consultas = Consulta.objects.select_related(
        'paciente', 'veterinario'
    ).prefetch_related('servicios').order_by('-fecha')
    context = {
        'consultas': consultas,
        'url_exportar': reverse('reportes:exportar_clinica_excel')
    }
    return render(request, 'reportes/clinica.html', context)


@admin_required
def exportar_clinica_excel(request):
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Clínica"
    
    # Encabezados
    headers = ['Fecha', 'Paciente', 'Tipo de Atención', 'Servicios', 'Veterinario']
    ws.append(headers)
    
    # Obtener datos
    consultas = Consulta.objects.select_related(
        'paciente', 'veterinario'
    ).prefetch_related('servicios').order_by('-fecha')
    
    # Agregar datos
    for consulta in consultas:
        fecha = consulta.fecha.strftime('%d/%m/%Y %H:%M') if consulta.fecha else '-'
        paciente = consulta.paciente.nombre if consulta.paciente else '-'
        tipo_atencion = consulta.get_tipo_consulta_display()
        servicios = consulta.servicios_nombres()
        veterinario = consulta.veterinario.get_full_name() if consulta.veterinario else '-'
        
        ws.append([
            fecha,
            paciente,
            tipo_atencion,
            servicios,
            veterinario
        ])
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_clinica.xlsx'
    
    # Guardar workbook en response
    wb.save(response)
    
    return response


@admin_required
def reporte_hospitalizacion(request):
    hospitalizaciones = Hospitalizacion.objects.select_related(
        'paciente', 'veterinario'
    ).order_by('-fecha_ingreso')
    
    context = {
        'hospitalizaciones': hospitalizaciones,
        'url_exportar': reverse('reportes:exportar_hospitalizacion_excel')
    }
    return render(request, 'reportes/hospitalizacion.html', context)


@admin_required
def exportar_hospitalizacion_excel(request):
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Hospitalización"
    
    # Encabezados
    headers = ['Fecha Ingreso', 'Fecha Alta', 'Paciente', 'Motivo', 'Estado', 'Veterinario']
    ws.append(headers)
    
    # Obtener datos
    hospitalizaciones = Hospitalizacion.objects.select_related(
        'paciente', 'veterinario'
    ).order_by('-fecha_ingreso')
    
    # Agregar datos
    for hosp in hospitalizaciones:
        fecha_ingreso = hosp.fecha_ingreso.strftime('%d/%m/%Y %H:%M') if hosp.fecha_ingreso else '-'
        fecha_alta = hosp.fecha_alta.strftime('%d/%m/%Y %H:%M') if hosp.fecha_alta else 'En curso'
        paciente = hosp.paciente.nombre if hosp.paciente else '-'
        motivo = hosp.motivo if hosp.motivo else '-'
        estado = hosp.get_estado_display()
        veterinario = hosp.veterinario.get_full_name() if hosp.veterinario else '-'
        
        ws.append([
            fecha_ingreso,
            fecha_alta,
            paciente,
            motivo,
            estado,
            veterinario
        ])
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_hospitalizacion.xlsx'
    
    # Guardar workbook en response
    wb.save(response)
    
    return response


@admin_required
def reporte_servicios(request):
    servicios_vendidos = DetalleVenta.objects.filter(
        tipo='servicio'
    ).select_related(
        'venta', 'venta__usuario_cobro', 'venta__usuario_creacion', 'servicio'
    ).order_by('-venta__fecha_creacion')
    
    context = {
        'servicios_vendidos': servicios_vendidos,
        'url_exportar': reverse('reportes:exportar_servicios_excel')
    }
    return render(request, 'reportes/servicios.html', context)


@admin_required
def exportar_servicios_excel(request):
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"
    
    # Encabezados
    headers = ['Fecha', 'Servicio', 'Precio', 'Estado', 'Usuario', 'Número Venta']
    ws.append(headers)
    
    # Obtener datos
    servicios_vendidos = DetalleVenta.objects.filter(
        tipo='servicio'
    ).select_related(
        'venta', 'venta__usuario_cobro', 'venta__usuario_creacion'
    ).order_by('-venta__fecha_creacion')
    
    # Agregar datos
    for detalle in servicios_vendidos:
        fecha = detalle.venta.fecha_creacion.strftime('%d/%m/%Y %H:%M') if detalle.venta.fecha_creacion else '-'
        usuario = detalle.venta.usuario_cobro.get_full_name() if detalle.venta.usuario_cobro else detalle.venta.usuario_creacion.get_full_name()
        precio = float(detalle.precio_unitario) if detalle.precio_unitario else 0.0
        estado = detalle.venta.get_estado_display()
        
        ws.append([
            fecha,
            detalle.descripcion,
            precio,
            estado,
            usuario,
            detalle.venta.numero_venta
        ])
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_servicios.xlsx'
    
    # Guardar workbook en response
    wb.save(response)
    
    return response
